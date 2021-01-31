import os
import sys
import psycopg2 as pg
sys.path.append(os.path.abspath("/home/suganesh/anaconda3/lib/python3.7/site-packages/"))
from django.shortcuts import render
import openpyxl
import pandas as pd
import numpy as np
import sqlalchemy
from sqlalchemy import create_engine
from django.http import HttpResponse,JsonResponse
from django.views.decorators.csrf import csrf_exempt
from rest_framework.parsers import JSONParser, MultiPartParser, FileUploadParser, FormParser
from rest_framework.exceptions import ParseError
from rest_framework import viewsets, permissions, generics, authentication, request, views
from rest_framework.response import Response
from rest_framework.views import APIView
from rest_framework_jwt import views as jwt_views
from django.http import HttpResponse
from tablib import Dataset
from . resources import EmployeeResource
from . models import (Employee, FileUpload, CreateDB, TableData, 
             FindDatatype, UserModel, UserMaster, WorkSpace)
from datasource.serializers import (FileUploadSerializer,CreateDBSerializer, TableDataSerializer,
    FindDataTypeSerializer, DatatypeSerializer, UserModelSerializer, 
    WorkSpaceSerializer,UserMasterSerializer, UserMasterCrtSerializer)
from rest_framework.generics import (ListCreateAPIView, ListAPIView, CreateAPIView)


class UserLoginViewJwt(jwt_views.ObtainJSONWebToken):
    serializer_class = UserModel

    def post(self, request, *args, **kwargs):
        response = super.post(request, *args, **kwargs)
        if response.status_code == status.HTTP_200_OK:
            user = get_user_model().objects.get(email=request.data[get_user_model().USERNAME_FIELD])
            serialized_user = self.user_serializer_class(user)
            response.data.update(serialized_user)
        return response

class UserModelListCrtApi(ListCreateAPIView):
    queryset = UserModel.objects.all()
    serializer_class = UserModelSerializer

    def get_queryset(self, *args, **kwargs):
        queryset_list = UserModel.objects.all()
        query = self.request.GET.get("EmpId")
        if query:
            queryset_list = queryset_list.filter(
                Q(EmpId__icontains=query)
            ).distinct()
        return queryset_list

class WorkSpaceView(ListCreateAPIView):
    queryset = WorkSpace.objects.all().values('WorkSpaceName')
    serializer_class = WorkSpaceSerializer
    print("working")

    def view_detail(self):
        query = self.request.GET.get('WorkSpaceName')
        for dbname in query['WorkSpaceName']:
            con = psycopg2.connect(user='postgres', host='127.0.0.1', password='sugan@123')
            con.set_isolation_level(ISOLATION_LEVEL_AUTOCOMMIT)
            cur = con.cursor()
            print("working")
            str = dbname
            cur.execute('CREATE DATABASE '+str)
            con.commit()
            cur.close()
            con.close()
            return HttpResponse(str)
        print("working")
# class WorkSpaceSerializerListCrtApi(ListCreateAPIView):
#     queryset = WorkSpace.objects.all()
#     serializer_class = WorkSpaceSerializer
#     def get_queryset(self):
#         # if request.method == 'GET':
#         WorkSpaceName = self.request.POST.get('WorkSpaceName')

#         # queryset_list = WorkSpace.objects.all()
#         # query = self.request.GET.get("WorkSpaceName")
#         # if query:
#         # dbname = "sample"
#         con = psycopg2.connect(user='postgres', host='127.0.0.1', password='sugan@123')
#         con.set_isolation_level(ISOLATION_LEVEL_AUTOCOMMIT)
#         cur = con.cursor()
#         str = "CREATE DATABASE {WorkSpaceName}".format(WorkSpaceName=WorkSpaceName)
#         cur.execute(sql.SQL(str))

#         # cur.execute('CREATE DATABASE  '+dbname)
#         # cur.flush()
#         con.commit()
#         cur.save()
#         cur.close()
#         con.close()
#         return HttpResponse


def home(request):
    print(request.GET)
    return render(request,"base.html")

def sample(request):
    return render(request,"sample.html")

def export_data(request):
    if request.method == 'POST':
        # Get selected option from form
        file_format = request.POST['file-format']
        employee_resource = EmployeeResource()
        dataset = employee_resource.export()
        if file_format == 'CSV':
            response = HttpResponse(dataset.csv, content_type='text/csv')
            response['Content-Disposition'] = 'attachment; filename="exported_data.csv"'
            return response
        elif file_format == 'JSON':
            response = HttpResponse(dataset.json, content_type='application/json')
            response['Content-Disposition'] = 'attachment; filename="exported_data.json"'
            return response
        elif file_format == 'XLS (Excel)':
            response = HttpResponse(dataset.xls, content_type='application/vnd.ms-excel')
            response['Content-Disposition'] = 'attachment; filename="exported_data.xls"'
            return response
    return render(request, 'export.html')

def import_data(request):
    if request.method == 'POST':
        file_format = request.POST['file-format']
        employee_resource = EmployeeResource()
        dataset = Dataset()
        try:
            new_employees = request.FILES['importData']
        except:
            raise

        if file_format == 'CSV':
            imported_data = dataset.load(new_employees.read().decode('utf-8'),format='csv')
            data = pd.read_csv(imported_data, index_col=0)
            datatype = data.dtypes
            # result = employee_resource.import_data(dataset, dry_run=True)

        elif file_format == 'JSON':
            imported_data = dataset.load(new_employees.read().decode('utf-8'),format='json')
            data = pd.read_json(imported_data, index_col=0)
            datatype = data.dtypes
            # result = employee_resource.import_data(dataset, dry_run=True)

        elif file_format == 'XLS (Excel)':
            imported_data = dataset.load(format='xlsx')
            data = pd.read_excel(imported_data, index_col=0)
            datatype = data.dtypes
            # result = employee_resource.import_data(dataset, dry_run=True)

        if not result.has_errors():
            employee_resource.import_data(dataset, dry_run=False)
    return render(request, 'import.html')

# from openpyxl import load_workbook
# filepath = "/home/suganesh/Downloads/testdata.xlsx"
# def get_sheetnames_xlsx(filepath, request):
#     if request.method == 'POST':
#         sheet_names = request.POST['sheet_names']
#         wb = load_workbook(filepath, read_only=True, keep_links=False)
#         return wb.sheetnames
#     else:
#         return "No file"
#     return render(request, 'export.html')

import psycopg2
from psycopg2 import sql
from psycopg2.extensions import ISOLATION_LEVEL_AUTOCOMMIT



# # @api_view(['GET','POST'])
# def dbcreation(request):
#     if request.method == 'POST':
#     # if "GET" == request.method:
#     #     dbname = request.POST.get('username')
#         return render(request, 'dbcreation.html', {})
#     else:
#         dbname = "sample"
#         con = psycopg2.connect(user='postgres', host='127.0.0.1',password='sugan@123')
#         con.set_isolation_level(ISOLATION_LEVEL_AUTOCOMMIT)
#         cur = con.cursor()
#         cur.execute(sql.SQL("CREATE DATABASE {}").format(sql.Identifier(dbname)))
#         # print("Your DB has been created!.. Your DB Name is :" + dbname + ";")
#         return render(request, 'dbcreation.html', {'username': dbname})


# ### File Upload Uisng Pandas ###
# import pandas as pd
# def handle_uploaded_file(f):
#     df = pd.read_excel(f, index_col=0)
#     file_DataType = df.dtypes

#  ### Image File upload ###
# class FileUploadView(views.APIView):
#     parser_classes = (FileUploadParser,)
#
#     def put(self, request, filename, format=None):
#         file_obj = request.FILES['file']
#         # do some stuff with uploaded file
#         return Response(status=204)


# class FileUploadView(views.APIView):
#     parser_classes = (MultiPartParser,FileUploadParser,FormParser,JSONParser)
#     queryset = FileUpload.objects.all()
#     serializer_class = FileUploadSerializer
#     http_method_names = ['get', 'head']
#
#     # def post(self, request, format = 'xlsx'):
#     #     file_obj = request.FILES['file']
#     #     return Response(status=204)
#
#     def pre_save(self,obj, format = None):
#         obj.file = self.request.FILES.get('file')
#     #
#     # def get(self,request, format = None):


class FileUploadListCrtApi(ListCreateAPIView):
    queryset = FileUpload.objects.all()
    serializer_class = FileUploadSerializer

    def get_queryset(self, *args, **kwargs):
        queryset_list = FileUpload.objects.all()
        query = self.request.FILES.get('file')
        # query = result.json()
        try:
            if query:
                for f in query:
                    if f.endswith(".xlsx"):
                        datatype = pd.read_excel(f)
                        data_type = datatype.dtypes
                        return data_type
                    elif f.endswith(".xls"):
                        datatype = pd.read_excel(f)
                        data_type = datatype.dtypes
                        return data_type
                    elif f.endswith(".xlsb"):
                        datatype = pd.read_excel(f)
                        data_type = datatype.dtypes
                        return data_type
                    elif f.endswith(".csv"):
                        datatype = pd.read_csv(f)
                        data_type = datatype.dtypes
                        return data_type
                    elif f.endswith(".json"):
                        datatype = pd.read_json(f)
                        data_type = datatype.dtypes
                        return data_type
                    elif f.endswith(".html"):
                        datatype = pd.read_html(f)
                        data_type = datatype.dtypes
                        return data_type
                    else:
                        return HttpResponse
                else:
                    return print("Please Contact to Administrator...!")
        except:
            raise
        if query:
            queryset_list = queryset_list.filter(
                Q(file__icontains=query)
            ).distinct()
        return queryset_list


class CreateDBListCrtApi(ListCreateAPIView):
    queryset = CreateDB.objects.all()
    serializer_class = CreateDBSerializer

    def dbcreation(self, *args, **kwargs):
        querylist = CreateDB.objects.all()
        query = self.request.GET.get('dbname')
        try:
            if query:
                con = psycopg2.connect(user='postgres', host='127.0.0.1', password='sugan@123')
                con.set_isolation_level(ISOLATION_LEVEL_AUTOCOMMIT)
                cur = con.cursor()
                cur.execute(sql.SQL("CREATE DATABASE {}").format(sql.Identifier(query)))
                print("Your DB has been created!.. Your DB Name is :" + query + ";")
                return
            else:
                print("DB Not Created, Please check with admin")
        except:
            raise
        if query:
            queryset_list = queryset_list.filter(
                Q(dbname_icontains = query)
            ).distinct()
        return querylist

import json
class TableDataListCrtApi(ListCreateAPIView):
    queryset = TableData.objects.all()
    serializer_class = TableDataSerializer

    def tabledata(self, *args, **kwargs):
        querylist = TableData.objects.all()
        query = self.request.GET.get('tableName')
        # result = self.response.json.get('tableName')
        if query:
            queryset_list = queryset_list.filter(
                Q(tableName__icontains = query)
            ).distinct()
        return querylist



class DatatypeListCrtApi(ListCreateAPIView):
    def datatype(self, *args, **kwargs):
        data = self.response.GET.get('dataTypeInfo')
        results = DatatypeSerializer(data, many=True).data
        return JsonResponse(results)
#
class FileUploadViewSet(viewsets.ViewSet):
    def create(self, request):
        serializer_class = FileSerializer(data=request.data)
        if 'file' not in request.FILES or not serializer_class.is_valid():
            return Response(status=status.HTTP_400_BAD_REQUEST)
        else:
            #Single File
            #handle_uploaded_file(request.FILES['file'])

            #Multiple Files
            files = request.FILES.getlist('file')
            for f in files:
                handle_uploaded_file(f)
            return Response(status=status.HTTP_201_CREATED)

class FindDataTypeListCrtApi(generics.ListCreateAPIView):
    queryset = FindDatatype.objects.all()
    serializer_class = FindDataTypeSerializer
    def find_datatype(self, *args, **kwargs):
        querylist = FileDatatype.objects.all()
        file_format = self.request.GET.get('filetype')
        file_dir = self.request.GET.get('filepath')
        file_name = self.request.GET.get('filename')
        HERE = os.path.abspath(os.path.dirname(__file__))
        DATA_DIR = os.path.abspath(os.path.join(HERE, file_dir))
        file_path = os.path.abspath(os.path.join(DATA_DIR, file_name))

    def post(self, request, *args, **kwargs):
        if file_format == 'Excel':
            xl = pd.ExcelFile(file_path)
            sheet_name = xl.sheet_names
            data = pd.read_excel(file_path)
            data_type = data.dtypes
            return(JsonResponse(data_type))

        elif file_format == 'CSV':
            data = pd.read_csv(file_path)
            data_type = data.dtypes
            return(JsonResponse(data_type))

        elif file_format == 'JSON':
            data = pd.read_json(file_path)
            data_type = data.dtypes
            return(JsonResponse(data_type))
        else:
            return(HttpResponse("Please Select Valid Format...!"))